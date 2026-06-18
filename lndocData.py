import logging
import re
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from scrapling import Selector

from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import Select, WebDriverWait

from session import session


LNDOC_URL = "https://tcfabprod.lexisnexis.com/ws/TrackerReport/iq/startQuery.jsp"
COURT_VALUE = "STTXPUC0"
TABLE_VALUE = "CSO_REC"
DEFAULT_DATE_FORMAT = "%m/%d/%Y"
FROM_LNDOC_HEADER = "From Ln-Doc"
SF_FILE_HEADER = "CSO_REC.CSO_SF_FILE"


# Examples handled:
#   ldc_smd_58153_3_1503136_e2e.pdf.00500000itfivr.pdf -> 58153_3_1503136
#   LDC_BC_58153_3_1503136.pdf -> 58153_3_1503136
LNDOC_FILENAME_PATTERN = re.compile(
    r"\bldc_(?:smd|bc)_(\d+_\d+_\d+)",
    re.IGNORECASE,
)

# Fallback for rows where the prefix is missing but the three numeric groups are present.
THREE_GROUP_PATTERN = re.compile(r"\b(\d+_\d+_\d+)\b")

SELECTED_COLUMNS = (
    ("CSO_REC.CSO_SID.VARCHAR2", "SID"),
    ("CSO_REC.CSOCRTJD_CPRCOURT.VARCHAR2", "Court"),
    ("CSO_REC.CSO_DT_ROUTED.DATE", "Date Routed"),
    ("CSO_REC.CSO_SF_FILE.VARCHAR2", "SF File"),
)


def lndoc_check(output_file, last_date=None, to_day_str=None):
    """
    Opens LN-DOC TrackerReport, builds a CSO_REC query for STTXPUC0, scrapes the
    results table with Scrapling, normalizes LN-DOC filenames, and writes the
    values into the existing Excel column named 'From Ln-Doc'.

    main.py can call either:
        lndoc_check(output_file)

    or, to use the GUI date range:
        lndoc_check(output_file, last_date, to_day_str)
    """
    driver = session()
    wait = WebDriverWait(driver, 600)
    date_ln = normalize_lndoc_date_range(last_date, to_day_str)

    try:
        logging.info("Opening LN-DOC browser")
        logging.info(f"Opening LN-DOC URL: {LNDOC_URL}")
        logging.info(f"LN-Doc Date Search: {date_ln}")

        driver.get(LNDOC_URL)
        wait_for_page_ready(driver, wait)

        start_new_query(driver, wait)
        select_cso_rec_table(driver, wait)
        select_result_columns(driver, wait)
        submit_search_filters(driver, wait, date_ln)

        has_results = wait_for_lndoc_results(driver)

        if not has_results:
            logging.info("No LN-DOC results found. Clearing From Ln-Doc column.")
            write_lndoc_results_to_excel(output_file, [])
            return

        html = get_best_results_html(driver)
        page = Selector(html)
        raw_file_names = scrape_lndoc_table_with_scrapling(page)

        formatted_results = format_lndoc_file_names(raw_file_names)
        write_lndoc_results_to_excel(output_file, formatted_results)

        logging.info(f"Extracted {len(raw_file_names)} LN-DOC document names.")
        logging.info(f"Unique formatted LN-DOC documents written: {len(formatted_results)}")
        logging.info("Document name formatting complete. Excel update successful.")

    except TimeoutException as error:
        logging.exception(f"LN-DOC check timed out: {error}")
        raise

    except Exception as error:
        logging.exception(f"LN-DOC check failed: {error}")
        raise

    finally:
        driver.quit()
        logging.info("LN-DOC browser closed.")


def normalize_lndoc_date_range(last_date=None, to_day_str=None):
    if last_date and to_day_str:
        return f"{last_date}-{to_day_str}"

    today = datetime.now().strftime(DEFAULT_DATE_FORMAT)
    logging.warning(
        "LN-DOC date range was not passed to lndoc_check; using today's date only. "
        "Pass lndoc_check(output_file, last_date, to_day_str) from main.py to use the GUI range."
    )
    return f"{today}-{today}"


def wait_for_page_ready(driver, wait):
    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")


def click_by_xpath(driver, wait, xpath, log_message):
    wait.until(ec.element_to_be_clickable((By.XPATH, xpath))).click()
    logging.info(log_message)


def start_new_query(driver, wait):
    click_by_xpath(
        driver,
        wait,
        '//input[@type="Radio" and @name="radOpenQuery" and @value="0"]',
        "Start a new query",
    )
    click_by_xpath(
        driver,
        wait,
        '//input[@type="Submit" and @name="startQuery" and @value="  Next  "]',
        "Clicked Next",
    )


def select_cso_rec_table(driver, wait):
    query_field = wait.until(ec.presence_of_element_located((By.NAME, "tableName")))
    Select(query_field).select_by_value(TABLE_VALUE)
    logging.info("Selected CSO_REC")

    click_by_xpath(
        driver,
        wait,
        '//input[@type="submit" and @value="  Next  "]',
        "Clicked Next",
    )


def select_result_columns(driver, wait):
    rec_field = wait.until(ec.presence_of_element_located((By.NAME, "selectedColumnNames")))
    select = Select(rec_field)

    try:
        select.deselect_all()
    except Exception:
        logging.debug("Could not deselect all selected columns; continuing.")

    for value, label in SELECTED_COLUMNS:
        select.select_by_value(value)
        logging.info(f"Select {label}")

    click_by_xpath(
        driver,
        wait,
        '//input[@type="submit" and @value="  Search  " and @onclick="goToSearch()"]',
        "Clicked Search setup",
    )


def submit_search_filters(driver, wait, date_ln):
    wait_for_page_ready(driver, wait)
    time.sleep(1)

    # Use the exact filter fields from the CSO_REC query page.
    # The row-label scanner was removed because it could grab the wrong input
    # and put STTXPUC0/date values into the wrong query fields.
    court_xpath = "/html/body/form/table/tbody/tr/td[2]/table[2]/tbody/tr[9]/td[4]/input"
    date_routed_xpath = "/html/body/form/table/tbody/tr/td[2]/table[2]/tbody/tr[11]/td[4]/input"
    submit_xpath = "/html/body/form/table/tbody/tr/td[2]/table[2]/tbody/tr[14]/td/input[5]"

    court_input = wait.until(ec.presence_of_element_located((By.XPATH, court_xpath)))
    court_input.clear()
    court_input.send_keys(COURT_VALUE)
    logging.info(f"Input Court: {COURT_VALUE}")

    date_input = wait.until(ec.presence_of_element_located((By.XPATH, date_routed_xpath)))
    date_input.clear()
    date_input.send_keys(date_ln)
    logging.info(f"Input Date Routed: {date_ln}")

    submit_button = wait.until(ec.element_to_be_clickable((By.XPATH, submit_xpath)))
    submit_button.click()
    logging.info("Submit Query")
    logging.info("Loading LN-DOC results.")


def find_filter_input_by_row_label(driver, label_text):
    label_text = label_text.lower()
    rows = driver.find_elements(By.CSS_SELECTOR, "form table tr")

    for row in rows:
        text = normalize_text(row.text).lower()
        if label_text not in text:
            continue

        inputs = row.find_elements(By.CSS_SELECTOR, "input[type='text'], input:not([type])")
        if inputs:
            return inputs[-1]

    return None


def find_submit_query_button(driver):
    submit_candidates = driver.find_elements(By.CSS_SELECTOR, "input[type='submit'], input[type='button']")

    for candidate in submit_candidates:
        value = normalize_text(candidate.get_attribute("value"))
        onclick = normalize_text(candidate.get_attribute("onclick"))
        if "submit" in value.lower() or "search" in value.lower() or "go" in onclick.lower():
            if candidate.is_displayed() and candidate.is_enabled():
                return candidate

    return None


def wait_for_lndoc_results(driver):
    try:
        WebDriverWait(driver, 600).until(
            lambda d: lndoc_has_result_table(d) or lndoc_has_no_results(d)
        )

        if lndoc_has_no_results(driver):
            logging.info("No results found on LN-DOC.")
            return False

        row_count = count_result_rows(driver)
        logging.info(f"LN-DOC result rows found: {row_count}")
        return row_count > 0

    except TimeoutException:
        logging.error("Timed out waiting for LN-DOC search results.")
        return False


def lndoc_has_result_table(driver):
    try:
        page_text = normalize_text(driver.find_element(By.TAG_NAME, "body").text).lower()

        # The generated LN-DOC result table uses DetailHeadText cells and places
        # filenames under the CSO_REC.CSO_SF_FILE header. Detect that exact
        # result shape instead of relying on generic table text.
        if SF_FILE_HEADER.lower() in page_text and ("ldc_smd" in page_text or "ldc_bc" in page_text):
            return True

        tables = driver.find_elements(By.CSS_SELECTOR, "table")
        for table in tables:
            text = normalize_text(table.text).lower()
            if SF_FILE_HEADER.lower() in text and ("ldc_smd" in text or "ldc_bc" in text):
                return True

        return False
    except Exception:
        return False


def lndoc_has_no_results(driver):
    try:
        page_text = normalize_text(driver.find_element(By.TAG_NAME, "body").text).lower()
        no_result_phrases = (
            "no records found",
            "no rows found",
            "no results found",
            "no data found",
            "returned no rows",
        )
        return any(phrase in page_text for phrase in no_result_phrases)
    except Exception:
        return False


def count_result_rows(driver):
    try:
        html = get_best_results_html(driver)
        page = Selector(html)
        return len(scrape_sf_file_column(page, log_rows=False))
    except Exception:
        return 0


def get_best_results_html(driver):
    driver.switch_to.default_content()
    return driver.page_source


def scrape_lndoc_table_with_scrapling(page):
    # Primary parser for the generated TrackerReport results table. The filename
    # is specifically under the CSO_REC.CSO_SF_FILE header, not necessarily in a
    # generic first/second column.
    raw_file_names = scrape_sf_file_column(page, log_rows=True)

    if raw_file_names:
        return raw_file_names

    logging.warning(
        "No LN-DOC filenames found under CSO_REC.CSO_SF_FILE. "
        "Running legacy row parser fallback."
    )

    rows = get_lndoc_result_rows(page)
    raw_file_names = []

    for row in rows:
        raw_name = extract_file_name_from_row(row)

        if not raw_name:
            continue

        raw_file_names.append(raw_name)

    if not raw_file_names:
        logging.warning("No LN-DOC filenames found from table cells. Running regex fallback.")
        raw_file_names = scrape_lndoc_filenames_by_regex(page)

    return raw_file_names


def scrape_sf_file_column(page, log_rows=False):
    raw_file_names = []
    table_info = find_sf_file_result_table(page)

    if table_info is None:
        return raw_file_names

    table, sf_file_index, header_row_index = table_info
    rows = table.css("tr")

    for row in rows[header_row_index + 1:]:
        cells = row.css("td")

        if len(cells) <= sf_file_index:
            continue

        raw_name = normalize_text(cells[sf_file_index].text)

        if not raw_name:
            continue

        # Skip footer/group rows and anything that is not an LN-DOC filename.
        if not LNDOC_FILENAME_PATTERN.search(raw_name):
            continue

        raw_file_names.append(raw_name)

    return raw_file_names


def find_sf_file_result_table(page):
    for table in page.css("table"):
        rows = table.css("tr")

        for row_index, row in enumerate(rows):
            cells = row.css("th, td")
            header_texts = [normalize_text(cell.text) for cell in cells]

            for col_index, header_text in enumerate(header_texts):
                if header_text.upper() == SF_FILE_HEADER:
                    logging.info(
                        f"Found LN-DOC filename column {SF_FILE_HEADER} "
                        f"at index {col_index}."
                    )
                    return table, col_index, row_index

    logging.warning(f"Could not find LN-DOC result header: {SF_FILE_HEADER}")
    return None


def get_lndoc_result_rows(page):
    candidate_tables = page.css("form table")

    if not candidate_tables:
        candidate_tables = page.css("table")

    best_rows = []

    for table in candidate_tables:
        table_text = normalize_text(table.text).lower()
        if not any(token in table_text for token in ("ldc_smd", "ldc_bc", COURT_VALUE.lower())):
            continue

        rows = table.css("tr")
        data_rows = []

        for row in rows:
            cells = row.css("td")
            if not cells:
                continue

            row_text = normalize_text(row.text).lower()
            if not row_text:
                continue

            # Skip filter/search-form rows. Result rows normally contain the court value or an LDC filename.
            if "date routed" in row_text and "sf file" in row_text:
                continue

            if "ldc_smd" in row_text or "ldc_bc" in row_text or COURT_VALUE.lower() in row_text:
                data_rows.append(row)

        if len(data_rows) > len(best_rows):
            best_rows = data_rows

    return best_rows


def extract_file_name_from_row(row):
    cells = row.css("td")

    # Prefer cells with LN-DOC filenames.
    for cell in cells:
        value = normalize_text(cell.text)
        if LNDOC_FILENAME_PATTERN.search(value):
            return value

        links = cell.css("a")
        for link in links:
            link_text = normalize_text(link.text)
            href = normalize_text(link.attrib.get("href", ""))
            if LNDOC_FILENAME_PATTERN.search(link_text):
                return link_text
            if LNDOC_FILENAME_PATTERN.search(href):
                return href

    # Preserve the old collector behavior as fallback: first non-empty cell.
    for cell in cells:
        value = normalize_text(cell.text)
        if value:
            return value

    return ""


def scrape_lndoc_filenames_by_regex(page):
    html_text = str(page.html_content)
    matches = []
    seen = set()

    for match in LNDOC_FILENAME_PATTERN.finditer(html_text):
        raw_value = match.group(0)
        if raw_value in seen:
            continue
        seen.add(raw_value)
        matches.append(raw_value)
        logging.info(f"Collected LN-DOC filename by regex fallback: {raw_value}")

    logging.info(f"Regex fallback LN-DOC filenames found: {len(matches)}")
    return matches


def format_lndoc_file_names(raw_file_names):
    formatted_results = []
    seen = set()

    for raw_name in raw_file_names:
        formatted = format_lndoc_file_name(raw_name)

        if not formatted:
            continue

        if formatted in seen:
            logging.info(f"Duplicate LN-DOC result skipped: {formatted}")
            continue

        seen.add(formatted)
        formatted_results.append(formatted)
        logging.info(f"{raw_name} -> Replaced: {formatted}")

    return formatted_results


def format_lndoc_file_name(raw_name):
    value = normalize_text(raw_name).lower()

    match = LNDOC_FILENAME_PATTERN.search(value)
    if match:
        return match.group(1)

    fallback_match = THREE_GROUP_PATTERN.search(value)
    if fallback_match:
        return fallback_match.group(1)

    return value


def write_lndoc_results_to_excel(output_file, formatted_results):
    output_file = Path(output_file)
    workbook = load_workbook(output_file)
    sheet = workbook.active

    headers = get_headers(sheet)

    if FROM_LNDOC_HEADER not in headers:
        raise ValueError(
            f"Excel file does not contain '{FROM_LNDOC_HEADER}' header. "
            "Create the workbook with headers: From Court, From Ln-Doc, Result, Link."
        )

    from_lndoc_col = headers[FROM_LNDOC_HEADER]

    clear_existing_column_values(sheet, from_lndoc_col)

    for row_number, formatted in enumerate(formatted_results, start=2):
        sheet.cell(row=row_number, column=from_lndoc_col).value = formatted

    auto_adjust_columns(sheet)
    workbook.save(output_file)

    logging.info(f"LN-DOC results written to Excel: {output_file}")


def get_headers(sheet):
    headers = {}

    for cell in sheet[1]:
        if cell.value:
            headers[normalize_text(cell.value)] = cell.column

    return headers


def clear_existing_column_values(sheet, column_number):
    for row_number in range(2, sheet.max_row + 1):
        sheet.cell(row=row_number, column=column_number).value = None


def auto_adjust_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = max_length + 5


def normalize_text(value):
    if value is None:
        return ""

    return " ".join(str(value).replace("\xa0", " ").split())
