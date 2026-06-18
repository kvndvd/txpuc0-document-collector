import logging

from openpyxl import load_workbook

from downloadFiles import download_included_file, get_download_filename


RESULT_DUPLICATE = "Duplicate"
RESULT_INCLUDE = "Include"
RESULT_DOWNLOAD_FAILED = "Download Failed"


def duplicate_check(output_file):
    logging.info("Duplicate check started.")

    workbook = load_workbook(output_file)
    sheet = workbook.active

    headers = get_headers(sheet)

    required_headers = ["From Court", "From Ln-Doc", "Result", "Link"]

    for header in required_headers:
        if header not in headers:
            raise ValueError(f"Excel file does not contain '{header}' header.")

    from_court_col = headers["From Court"]
    from_lndoc_col = headers["From Ln-Doc"]
    result_col = headers["Result"]
    link_col = headers["Link"]

    lndoc_values = get_column_values(sheet, from_lndoc_col)

    include_count = 0
    duplicate_count = 0
    download_count = 0
    download_failed_count = 0

    for row_number in range(2, sheet.max_row + 1):
        court_value = normalize_text(
            sheet.cell(row=row_number, column=from_court_col).value
        )

        if not court_value:
            continue

        if court_value in lndoc_values:
            sheet.cell(row=row_number, column=result_col).value = RESULT_DUPLICATE
            duplicate_count += 1
            logging.info(f"{court_value}: {RESULT_DUPLICATE}")
            continue

        original_pdf_link = normalize_text(
            sheet.cell(row=row_number, column=link_col).value
        )

        if download_included_file(output_file, court_value, original_pdf_link):
            formatted_filename = get_download_filename(court_value)
            sheet.cell(row=row_number, column=result_col).value = RESULT_INCLUDE
            sheet.cell(row=row_number, column=link_col).value = formatted_filename
            include_count += 1
            download_count += 1
            logging.info(f"{RESULT_INCLUDE} - {formatted_filename}")
        else:
            sheet.cell(row=row_number, column=result_col).value = RESULT_DOWNLOAD_FAILED
            sheet.cell(row=row_number, column=link_col).value = original_pdf_link
            download_failed_count += 1
            logging.info(f"{court_value}: {RESULT_DOWNLOAD_FAILED} | Check court link--")

    workbook.save(output_file)

    logging.info("Duplicate check completed.")
    logging.info(f"Duplicate count: {duplicate_count}")
    logging.info(f"Include count: {include_count}")
    logging.info(f"Downloaded count: {download_count}")
    logging.info(f"Download failed count: {download_failed_count}")

    return duplicate_count, include_count, download_failed_count


def get_headers(sheet):
    headers = {}

    for cell in sheet[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    return headers


def get_column_values(sheet, column_number):
    values = set()

    for row_number in range(2, sheet.max_row + 1):
        value = normalize_text(sheet.cell(row=row_number, column=column_number).value)

        if value:
            values.add(value)

    return values


def normalize_text(value):
    if value is None:
        return ""

    return str(value).replace("\xa0", " ").strip()
